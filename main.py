#!/usr/bin/env python

import os
import sys
import csv
from typing import List

import pandas as pd
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

from service.code39 import generate_code39_images


# from PIL import Image
# from reportlab.pdfgen import canvas
# from reportlab.lib.pagesizes import A4
  
# def generate_concat_pdf(image_paths, output_pdf_path, images_per_row=3, images_per_column=4, padding=15):
#   c = canvas.Canvas(output_pdf_path, pagesize=A4)
#   width, height = A4
#   image_width = (width - (images_per_row + 1) * padding) / images_per_row
#   image_height = (height - (images_per_column + 1) * padding) / images_per_column

#   for i, image_path in enumerate(image_paths):
#     if i % (images_per_row * images_per_column) == 0 and i > 0:
#       c.showPage()
#     img = Image.open(image_path)
#     img.thumbnail((image_width, image_height))
#     x = padding + (i % images_per_row) * (image_width + padding)
#     y = height - ((i // images_per_row) % images_per_column + 1) * (image_height + padding)
#     c.drawInlineImage(img, x, y, width=img.width, height=img.height)

#   c.save()

def excel_to_csv(xlsx_path: str, csv_path: str) -> None:
  df = pd.read_excel(xlsx_path, header=None)
  os.remove(csv_path) if os.path.exists(csv_path) else None
  df.to_csv(csv_path, index=False, header=False)


def create_excel_with_images(images_dir, images_per_row, images_per_col, gap):
  # 新規ワークブックを作成
  wb = Workbook()

  # 画像ファイルのリストを取得
  image_files = [f for f in os.listdir(images_dir) if f.endswith('.png')]

  # A4サイズの印刷範囲（ピクセル単位）
  a4_width = 21 * 37.795275591  # cm to pixel conversion
  a4_height = 29.7 * 37.795275591  # cm to pixel conversion

  # 各画像のサイズを計算
  image_width = (a4_width - (images_per_row + 1) * gap) // images_per_row
  image_height = (a4_height - (images_per_col + 1) * gap) // images_per_col

  # 各画像を貼り付ける位置を計算
  positions = [(row, col) for row in range(images_per_row) for col in range(images_per_col)]

  # 各画像を貼り付け
  for i, image_file in enumerate(image_files):
    # 新しいシートを作成する必要があるかどうかを確認
    if i != 0 and i % (images_per_row * images_per_col) == 0:
      ws = wb.create_sheet()

    ws = wb.worksheets[i // (images_per_row * images_per_col)]

    # 画像を開き、サイズを変更（縦横比維持）
    img = Image.open(os.path.join(images_dir, image_file))
    img.thumbnail((image_width, image_height))

    # PIL Imageからopenpyxl Imageに変換
    img.save('temp.png')
    xl_img = XLImage('temp.png')

    # 指定された位置に画像を貼り付け
    row, col = positions[i % (images_per_row * images_per_col)]
    ws.add_image(xl_img, ws.cell(row=row+1, column=col+1).coordinate)

    # セルの高さと幅を調整（画像と同じサイズに）
    ws.row_dimensions[row+1].height = image_height
    ws.column_dimensions[chr(col+65)].width = image_width / 6  # Excelの列幅は文字数で指定

  # ワークブックを保存
  wb.save('output.xlsx')


if __name__ == "__main__":
  args = sys.argv
  xlsx_path = args[1]
  dest_dir = args[2]

  excel_to_csv(xlsx_path=xlsx_path, csv_path="./tmp.csv")
  code39_images_path = generate_code39_images(csv_path="./tmp.csv", dest_dir=dest_dir)
  create_excel_with_images(images_dir=dest_dir, images_per_row=4, images_per_col=4, gap=0)
  # generate_concat_pdf(image_paths=code39_images_path, output_pdf_path="./concat.pdf")
